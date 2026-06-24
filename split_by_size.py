import os
import math
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

# Put the name of your massive downloaded file here
input_file = r"C:\Users\BGVK38\Downloads\DRY RUN 5_6_2026.xlsx"
max_mb = 80.0

file_size_mb = os.path.getsize(input_file) / (1024 * 1024)
print(f"Original file size: {file_size_mb:.2f} MB")

print("Loading original workbook to extract images safely... (This might take a minute)")
wb_in = load_workbook(input_file)
ws_in = wb_in.active

# 1. ULTIMATE IMAGE RADAR: Finds 100% of images no matter how Excel formatted them
print("Extracting images into memory...")
image_dict = {}
for img in ws_in._images:
    orig_row = None
    try:
        if hasattr(img, 'anchor'):
            if hasattr(img.anchor, '_from'):
                orig_row = img.anchor._from.row + 1
            elif hasattr(img.anchor, 'row'):
                orig_row = img.anchor.row + 1
            elif isinstance(img.anchor, str):
                import re
                m = re.search(r'\d+', img.anchor)
                if m: orig_row = int(m.group())
    except Exception: pass
    
    if orig_row is not None:
        img_bytes = None
        try:
            if hasattr(img, 'ref') and hasattr(img.ref, 'read'):
                img.ref.seek(0) 
                img_bytes = img.ref.read()
            elif hasattr(img, '_data'):
                data = img._data() if callable(img._data) else img._data
                if hasattr(data, 'read'):
                    data.seek(0)
                    img_bytes = data.read()
                elif isinstance(data, bytes):
                    img_bytes = data
        except Exception as e:
            print(f"Warning: Failed to read image at row {orig_row}: {e}")
            
        if img_bytes:
            image_dict[orig_row] = {
                "bytes": img_bytes,
                "width": img.width,
                "height": img.height
            }

print(f"Successfully extracted {len(image_dict)} images!")

max_row = ws_in.max_row
last_row_val = str(ws_in.cell(row=max_row, column=1).value or "").strip()
has_summary = "SUMMARY" in last_row_val.upper()

total_data_rows = max_row - 1 
if has_summary:
    total_data_rows -= 1

# -------------------------------------------------------------------------
# SMART DUPLICATE SCANNER (Index + Device + Language Lock)
# -------------------------------------------------------------------------
print("\nScanning for extra duplicate test runs...")

headers = [str(ws_in.cell(row=1, column=c).value).strip().lower() for c in range(1, ws_in.max_column + 1)]
dev_col = headers.index("device") + 1 if "device" in headers else 2
lang_col = headers.index("language") + 1 if "language" in headers else 4
idx_col = headers.index("index") + 1 if "index" in headers else 7

latest_row_map = {}
rows_with_no_index = []

for r in range(2, total_data_rows + 2):
    idx_val = str(ws_in.cell(row=r, column=idx_col).value or "").strip()
    dev_val = str(ws_in.cell(row=r, column=dev_col).value or "").strip()
    lang_val = str(ws_in.cell(row=r, column=lang_col).value or "").strip()
    
    if idx_val and idx_val.lower() != 'nan' and idx_val != 'none' and "skip" not in idx_val.lower():
        unique_key = (idx_val, dev_val, lang_val)
        latest_row_map[unique_key] = r
    else:
        rows_with_no_index.append(r)

rows_to_keep_globally = sorted(list(latest_row_map.values()) + rows_with_no_index)
removed_count = total_data_rows - len(rows_to_keep_globally)
print(f"Identified {removed_count} actual duplicate rows (crashed/retried tests) to remove.")

estimated_new_mb = file_size_mb * (len(rows_to_keep_globally) / float(max(1, total_data_rows)))
print(f"Estimated file size AFTER removing duplicates: {estimated_new_mb:.2f} MB")

if estimated_new_mb <= max_mb:
    num_files_needed = 1
    kept_rows_per_file = len(rows_to_keep_globally)
    print(f"The cleaned file is under {max_mb} MB. Generating ONE cleaned file!")
else:
    num_files_needed = math.ceil(estimated_new_mb / max_mb)
    kept_rows_per_file = math.ceil(len(rows_to_keep_globally) / num_files_needed)
    print(f"Splitting {len(rows_to_keep_globally)} kept rows into {num_files_needed} files (approx. {kept_rows_per_file} rows each)...")
wb_in.close()

# 2. Build the perfectly formatted files
for i in range(num_files_needed):
    chunk_kept_orig_rows = rows_to_keep_globally[i * kept_rows_per_file : (i+1) * kept_rows_per_file]
    chunk_kept_set = set(chunk_kept_orig_rows)
    
    print(f"\nProcessing Part {i+1}...")
    
    wb_part = load_workbook(input_file)
    ws_part = wb_part.active
    ws_part._images = []
    
    if has_summary:
        ws_part.delete_rows(ws_part.max_row)
        
    rows_to_delete = sorted([r for r in range(2, total_data_rows + 2) if r not in chunk_kept_set])
    
    print(f" -> Deleting {len(rows_to_delete)} unneeded rows from this copy (in bulk)...")
    
    # -------------------------------------------------------------------------
    # CRITICAL PERFORMANCE FIX: Bulk Deletion Algorithm
    # -------------------------------------------------------------------------
    if rows_to_delete:
        blocks = []
        start_r = rows_to_delete[0]
        end_r = rows_to_delete[0]
        
        # Group single rows into massive contiguous blocks
        for r in rows_to_delete[1:]:
            if r == end_r + 1:
                end_r = r
            else:
                blocks.append((start_r, end_r))
                start_r = r
                end_r = r
        blocks.append((start_r, end_r))
        
        # Delete massive blocks from bottom to top so row numbers don't shift!
        for start_b, end_b in reversed(blocks):
            amount = end_b - start_b + 1
            ws_part.delete_rows(start_b, amount)
    # -------------------------------------------------------------------------
        
    print(f" -> Repasting images perfectly into the cleaned sheet...")
    current_row = 2
    for orig_r in chunk_kept_orig_rows:
        if orig_r in image_dict:
            img_info = image_dict[orig_r]
            try:
                new_img = OpenpyxlImage(io.BytesIO(img_info["bytes"]))
                new_img.width = img_info["width"]
                new_img.height = img_info["height"]
                
                col_idx = 13 
                r_idx = current_row - 1 
                
                marker_from = AnchorMarker(col=col_idx, colOff=0, row=r_idx, rowOff=0)
                marker_to = AnchorMarker(col=col_idx + 1, colOff=0, row=r_idx + 1, rowOff=0)
                new_img.anchor = TwoCellAnchor(editAs='twoCell', _from=marker_from, to=marker_to)
                
                ws_part.add_image(new_img)
            except Exception as e:
                print(f"    Warning: Failed to restore image on row {current_row}: {e}")
                
        current_row += 1
        
    p = f = w = s = 0
    for r in range(2, ws_part.max_row + 1):
        verdict = str(ws_part.cell(row=r, column=12).value or "").strip().upper()
        if verdict == "PASS": p += 1
        elif verdict == "FAIL": f += 1
        elif verdict == "WARN": w += 1
        elif verdict == "SKIP": s += 1
        
    processed_count = p + f + w
    est_seconds = processed_count * 36.75
    mins = int(est_seconds // 60)
    secs = est_seconds % 60
    
    summary_row = [
        f"SUMMARY (Part {i+1})" if num_files_needed > 1 else "SUMMARY", 
        f"Total Time: ~{mins}m {secs:.0f}s", 
        f"PASS: {p}", 
        f"FAIL: {f}", 
        f"WARN: {w}", 
        f"SKIP: {s}"
    ]
    
    ws_part.append(summary_row)
    
    new_max_row = ws_part.max_row
    ws_part.row_dimensions[new_max_row].height = 25
    
    summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    summary_font = Font(bold=True)
    for col_num in range(1, 15):
        cell = ws_part.cell(row=new_max_row, column=col_num)
        cell.fill = summary_fill
        cell.font = summary_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    output_name = f"Batch_Summary_Cleaned.xlsx" if num_files_needed == 1 else f"Batch_Summary_Part_{i+1}.xlsx"
    wb_part.save(output_name)
    wb_part.close()
    print(f"-> Saved {output_name} successfully!")

print("\nDone! The extra duplicates were deleted and images are securely locked to the cells.")